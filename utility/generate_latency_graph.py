"""
Module used to:
1. Collect the latencies from the provided logs,
2. generates a graph for the top N latencies against time

The input file is a modified file, which contains  latencies are grepped from ceph OSD log for faster run.
-> grep -a dequeue_op.*latency ceph-osd.log > latencies.log
( Performed to reduce run time, since lots of logging would be generated )

NOTE : The debug levels for the OSD for the latencies to be captured,
        the Param: debug_osd needs to be set to 20 globally
"""

import re
import sys
from datetime import datetime

import openpyxl
from docopt import docopt
from openpyxl.chart import LineChart, Reference

doc = """
Usage:
  generate_latency_graph.py --latencies-log <file_name> [--num-ops <ops>]

Options:
  --latencies-log <name>                Name of the file, where the latencies are grepped from ceph OSD log
                                        -> grep -a dequeue_op.*latency ceph-osd.log > latencies.log
  ----num-ops <num>                     Top N latencies to be collected for the graph
"""


class Op(object):
    def __init__(self, line):
        self.lat = float(re.search(r"latency (\d+.\\d+)", line).group(1))
        self.line = line
        time_format = "%Y-%m-%d %H:%M:%S.%f"
        self.time = datetime.strptime(line[:23], time_format)


def run(args):
    logfile = args["--latencies-log"]
    num_ops = int(args.get("--num-ops", 10000))
    ops = []
    top_ops = []

    with open(logfile) as f:
        for line in f.readlines():
            ops.append(Op(line))
    ops.sort(key=lambda x: x.time)
    print(f"{num_ops} longest requests:")
    for i in range(min(num_ops, len(ops))):
        top_ops.append((ops[i].lat, ops[i].time, ops[i].line))
    top_ops.sort(key=lambda x: x[1])
    # Creating workbook and adding data to it
    wb = openpyxl.Workbook()
    sheet = wb.active
    text = f"Operations from Timestamp :{ops[0].time} - {ops[-1].time}"
    sheet.append([text])
    for val in top_ops:
        sheet.append([val[0], val[1]])

    max_row = sheet.max_row  # if sheet.max_row < 100 else 100
    values = Reference(sheet, min_col=1, min_row=2, max_col=1, max_row=max_row)
    # Create object of LineChart class
    chart = LineChart()
    chart.add_data(values)
    # set the title of the chart
    chart.title = f" Latencies {text}"
    chart.x_axis.title = " Time progression ----> "
    chart.y_axis.title = " Time in sec "
    sheet.add_chart(chart, "E2")

    # save the file
    fname = f"{sys.argv[1]}.xlsx"
    wb.save(fname)
    print("Done!")


if __name__ == "__main__":
    args = docopt(doc)
    try:
        run(args)
    except Exception as err:
        print(
            f"Exception hit while generating graph for the latencies provided.\n error : {err} "
        )
        exit(1)
